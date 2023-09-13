from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title = "Capital")

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ["KA", "TA", "TN"]

sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row :
        cell.font = ft

for x in range(2, 5):
    sheet.cell(row = x, column = 1).value = state[x-2]
    sheet.cell(row = x, column = 2).value = lang[x-2]
    sheet.cell(row = x, column = 3).value = code[x-2]

wb.save("Excel.xlsx")

key = input("Enter the state code to display the language")
for x in range(2,5):
    data = sheet.cell(row=x, column=3).value
    if data  == key:
        print("Corresponding language is ", sheet.cell(row=x,column=2).value)
    